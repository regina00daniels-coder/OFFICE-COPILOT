from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, RequestFactory, TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.tenants.models import Tenant
from .models import Task, TaskAnalysisRun
from .views import task_list_page


class TaskExcelTests(TestCase):
    def setUp(self):
        self.tenant_a = Tenant.objects.create(name="Tenant A", domain="a.local")
        self.tenant_b = Tenant.objects.create(name="Tenant B", domain="b.local")

        user_model = get_user_model()
        self.user_a = user_model.objects.create_user(
            username="alice",
            password="pass1234",
            tenant=self.tenant_a,
            role=user_model.Role.ADMIN,
        )
        self.user_b = user_model.objects.create_user(
            username="bob",
            password="pass1234",
            tenant=self.tenant_b,
            role=user_model.Role.USER,
        )
        self.assignee_a = user_model.objects.create_user(
            username="assignee_a",
            password="pass1234",
            tenant=self.tenant_a,
            role=user_model.Role.USER,
        )

        Task.objects.create(
            tenant=self.tenant_a,
            title="Tenant A task",
            status=Task.Status.TODO,
            priority=Task.Priority.HIGH,
            due_date=date(2026, 2, 20),
            created_by=self.user_a,
            assigned_to=self.assignee_a,
        )
        Task.objects.create(
            tenant=self.tenant_b,
            title="Tenant B task",
            status=Task.Status.DONE,
            priority=Task.Priority.LOW,
            created_by=self.user_b,
        )

        self.client = Client()
        self.factory = RequestFactory()

    def _build_import_file(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["title", "description", "status", "priority", "due_date", "assigned_to"])
        for row in rows:
            worksheet.append(row)
        content = BytesIO()
        workbook.save(content)
        return SimpleUploadedFile(
            "tasks.xlsx",
            content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_export_excel_is_tenant_scoped(self):
        self.client.login(username="alice", password="pass1234")
        response = self.client.get(reverse("task-export-excel"), HTTP_X_TENANT="a.local", HTTP_HOST="localhost")

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        self.assertEqual(rows[0][1], "Title")
        exported_titles = {r[1] for r in rows[1:]}
        self.assertIn("Tenant A task", exported_titles)
        self.assertNotIn("Tenant B task", exported_titles)

    def test_task_list_page_renders(self):
        request = self.factory.get(reverse("task-list-page"), HTTP_X_TENANT="a.local", HTTP_HOST="localhost")
        request.user = self.user_a
        request.tenant = self.tenant_a
        request.session = {}
        response = task_list_page(request)
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Task Workspace", response.content)
        self.assertIn(b"Tenant A task", response.content)

    def test_analyst_run_generates_workbook(self):
        analyst_file = self._build_import_file(
            [
                ["Data task 1", "desc", "todo", "high", "2026-02-25", "assignee_a"],
                ["Data task 2", "desc", "in_progress", "low", "2026-03-10", "assignee_a"],
            ]
        )
        self.client.login(username="alice", password="pass1234")
        response = self.client.post(
            reverse("task-analyst-run"),
            data={"file": analyst_file},
            HTTP_X_TENANT="a.local",
            HTTP_HOST="localhost",
        )
        self.assertEqual(response.status_code, 302)
        run = TaskAnalysisRun.objects.filter(tenant=self.tenant_a).first()
        self.assertIsNotNone(run)
        self.assertEqual(run.status, TaskAnalysisRun.Status.COMPLETED)
        self.assertTrue(bool(run.workbook_file))

    def test_import_excel_creates_valid_rows_and_reports_errors(self):
        upload = self._build_import_file(
            [
                ["Imported valid", "new task", "todo", "medium", "2026-02-25", "assignee_a"],
                ["Bad status row", "invalid status", "blocked", "low", "2026-02-25", "assignee_a"],
                ["Cross-tenant assignee", "invalid assignee", "todo", "high", "2026-02-26", "bob"],
                ["", "missing title", "todo", "high", "2026-02-26", "assignee_a"],
            ]
        )
        self.client.login(username="alice", password="pass1234")
        response = self.client.post(
            reverse("task-import-excel"),
            data={"file": upload},
            HTTP_X_TENANT="a.local",
            HTTP_HOST="localhost",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["total_rows_processed"], 4)
        self.assertEqual(payload["rows_inserted"], 1)
        self.assertEqual(payload["rows_skipped"], 3)
        self.assertEqual(len(payload["errors"]), 3)
        self.assertTrue(Task.objects.filter(tenant=self.tenant_a, title="Imported valid").exists())
        self.assertFalse(Task.objects.filter(tenant=self.tenant_a, title="Bad status row").exists())
