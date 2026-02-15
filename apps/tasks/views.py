import json
from datetime import date, datetime
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from office_copilot.authz import enforce_tenant_access
from apps.accounts.models import User
from apps.reporting.models import Report
from .analytics import analyze_task_dataframe
from .models import Task, TaskAnalysisRun


@login_required
@require_http_methods(["GET", "POST"])
def task_list_create(request):
    enforce_tenant_access(request)

    if request.method == "GET":
        tasks = Task.objects.filter(tenant=request.tenant).select_related("assigned_to")
        return JsonResponse(
            {
                "results": [
                    {
                        "id": t.id,
                        "title": t.title,
                        "status": t.status,
                        "priority": t.priority,
                        "assigned_to": t.assigned_to.username if t.assigned_to else None,
                    }
                    for t in tasks
                ]
            }
        )

    payload = json.loads(request.body or "{}")
    task = Task.objects.create(
        tenant=request.tenant,
        title=payload["title"],
        description=payload.get("description", ""),
        status=payload.get("status", Task.Status.TODO),
        priority=payload.get("priority", Task.Priority.MEDIUM),
        created_by=request.user,
    )
    return JsonResponse({"id": task.id, "title": task.title}, status=201)


def _parse_iso_date(value: str, field_name: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ValidationError(f"{field_name} must be YYYY-MM-DD") from exc


def _parse_import_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    raise ValidationError("due_date must be a valid date")


def _normalize_header(value: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in str(value).strip().lower()).strip("_")


def _import_tasks_from_upload(request, upload, mapping_payload=None):
    try:
        workbook = load_workbook(upload, data_only=True)
    except Exception:
        return 400, {"detail": "Invalid Excel file"}

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return 400, {"detail": "Excel file is empty"}

    header_row = rows[0]
    if not any(header_row):
        return 400, {"detail": "Header row is empty"}

    header_index = {_normalize_header(value): idx for idx, value in enumerate(header_row) if value is not None}
    mapping = {
        "title": "title",
        "description": "description",
        "status": "status",
        "priority": "priority",
        "due_date": "due_date",
        "assigned_to": "assigned_to",
    }
    if mapping_payload:
        try:
            custom_mapping = json.loads(mapping_payload)
        except json.JSONDecodeError:
            return 400, {"detail": "Invalid mapping JSON"}
        for destination, source in custom_mapping.items():
            mapping[_normalize_header(destination)] = _normalize_header(source)

    field_indexes = {}
    for field, source_header in mapping.items():
        if source_header in header_index:
            field_indexes[field] = header_index[source_header]

    if "title" not in field_indexes:
        return 400, {"detail": "Could not map required column: title"}

    processed = 0
    created = 0
    skipped = 0
    errors = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        processed += 1
        try:
            title_value = row[field_indexes["title"]]
            title = str(title_value).strip() if title_value not in (None, "") else ""
            if not title:
                raise ValidationError("title is required")

            description = ""
            if "description" in field_indexes and field_indexes["description"] < len(row):
                description_value = row[field_indexes["description"]]
                description = str(description_value).strip() if description_value not in (None, "") else ""

            status = Task.Status.TODO
            if "status" in field_indexes and field_indexes["status"] < len(row):
                status_value = row[field_indexes["status"]]
                if status_value not in (None, ""):
                    status = str(status_value).strip().lower()
            if status not in Task.Status.values:
                raise ValidationError("status must be one of: todo, in_progress, done")

            priority = Task.Priority.MEDIUM
            if "priority" in field_indexes and field_indexes["priority"] < len(row):
                priority_value = row[field_indexes["priority"]]
                if priority_value not in (None, ""):
                    priority = str(priority_value).strip().lower()
            if priority not in Task.Priority.values:
                raise ValidationError("priority must be one of: low, medium, high")

            due_date = None
            if "due_date" in field_indexes and field_indexes["due_date"] < len(row):
                due_date = _parse_import_date(row[field_indexes["due_date"]])

            assigned_to = None
            if "assigned_to" in field_indexes and field_indexes["assigned_to"] < len(row):
                assigned_to_value = row[field_indexes["assigned_to"]]
                if assigned_to_value not in (None, ""):
                    assigned_to = User.objects.filter(
                        tenant=request.tenant,
                        username=str(assigned_to_value).strip(),
                    ).first()
                    if assigned_to is None:
                        raise ValidationError("assigned_to user not found in tenant")

            Task.objects.create(
                tenant=request.tenant,
                title=title,
                description=description,
                status=status,
                priority=priority,
                due_date=due_date,
                assigned_to=assigned_to,
                created_by=request.user,
            )
            created += 1
        except ValidationError as exc:
            skipped += 1
            errors.append({"row": row_number, "error": exc.message})
        except Exception as exc:
            skipped += 1
            errors.append({"row": row_number, "error": str(exc)})

    return 200, {
        "total_rows_processed": processed,
        "rows_inserted": created,
        "rows_skipped": skipped,
        "errors": errors,
    }


@login_required
@require_http_methods(["GET"])
def task_list_page(request):
    enforce_tenant_access(request)
    tasks = Task.objects.filter(tenant=request.tenant).select_related("assigned_to").order_by("due_date", "-created_at")
    context = {
        "tasks": tasks,
        "import_result": request.session.pop("task_import_result", None),
        "status_choices": Task.Status.choices,
        "priority_choices": Task.Priority.choices,
    }
    return render(request, "tasks/list.html", context)


@login_required
@require_http_methods(["POST"])
def task_import_page(request):
    enforce_tenant_access(request)
    upload = request.FILES.get("file")
    if not upload:
        request.session["task_import_result"] = {"detail": "Missing file upload"}
        return redirect("task-list-page")

    status_code, payload = _import_tasks_from_upload(request, upload, request.POST.get("mapping"))
    request.session["task_import_result"] = payload
    if status_code != 200:
        request.session["task_import_result"] = {"detail": payload.get("detail", "Import failed")}
    return redirect("task-list-page")


@login_required
@require_http_methods(["GET"])
def task_analyst_page(request):
    enforce_tenant_access(request)
    runs = TaskAnalysisRun.objects.filter(tenant=request.tenant).select_related("user")[:20]
    context = {
        "runs": runs,
        "analyst_result": request.session.pop("analyst_result", None),
    }
    return render(request, "tasks/analyst.html", context)


@login_required
@require_http_methods(["POST"])
def task_analyst_run(request):
    enforce_tenant_access(request)
    upload = request.FILES.get("file")
    if not upload:
        request.session["analyst_result"] = {"detail": "Missing Excel file upload"}
        return redirect("task-analyst-page")

    run = TaskAnalysisRun.objects.create(
        tenant=request.tenant,
        user=request.user,
        source_file=upload,
        status=TaskAnalysisRun.Status.PROCESSING,
    )
    try:
        run.source_file.open("rb")
        summary, workbook_bytes = analyze_task_dataframe(run.source_file)
        run.source_file.close()
        filename = f"task_analytics_{run.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        run.workbook_file.save(filename, ContentFile(workbook_bytes), save=False)
        run.summary = summary
        run.status = TaskAnalysisRun.Status.COMPLETED
        run.save(update_fields=["workbook_file", "summary", "status"])
        Report.objects.create(
            tenant=request.tenant,
            name=f"Task Analyst Report {run.id}",
            report_type=Report.Type.TASKS,
            payload={"analysis_run_id": run.id, **summary},
            generated_by=request.user,
        )
        request.session["analyst_result"] = {
            "run_id": run.id,
            "rows_uploaded": summary["rows_uploaded"],
            "rows_after_cleaning": summary["rows_after_cleaning"],
            "rows_removed": summary["rows_removed"],
            "duplicate_titles": summary["duplicate_titles"],
            "anomalous_due_dates": summary["anomalous_due_dates"],
        }
    except Exception as exc:
        run.status = TaskAnalysisRun.Status.FAILED
        run.summary = {"error": str(exc)}
        run.save(update_fields=["status", "summary"])
        request.session["analyst_result"] = {"detail": str(exc)}
    return redirect("task-analyst-page")


@login_required
@require_http_methods(["GET"])
def task_analyst_download(request, run_id):
    enforce_tenant_access(request)
    run = get_object_or_404(TaskAnalysisRun, id=run_id, tenant=request.tenant)
    if not run.workbook_file:
        return JsonResponse({"detail": "Workbook not available for this run"}, status=404)
    run.workbook_file.open("rb")
    data = run.workbook_file.read()
    run.workbook_file.close()
    response = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    name = run.workbook_file.name.split("/")[-1]
    response["Content-Disposition"] = f'attachment; filename="{name}"'
    return response


@login_required
@require_http_methods(["GET"])
def task_export_excel(request):
    enforce_tenant_access(request)
    tasks = Task.objects.filter(tenant=request.tenant).select_related("assigned_to", "created_by")

    status = request.GET.get("status")
    if status:
        if status not in Task.Status.values:
            return JsonResponse({"detail": "Invalid status filter"}, status=400)
        tasks = tasks.filter(status=status)

    priority = request.GET.get("priority")
    if priority:
        if priority not in Task.Priority.values:
            return JsonResponse({"detail": "Invalid priority filter"}, status=400)
        tasks = tasks.filter(priority=priority)

    due_from = request.GET.get("due_from")
    if due_from:
        try:
            tasks = tasks.filter(due_date__gte=_parse_iso_date(due_from, "due_from"))
        except ValidationError as exc:
            return JsonResponse({"detail": exc.message}, status=400)

    due_to = request.GET.get("due_to")
    if due_to:
        try:
            tasks = tasks.filter(due_date__lte=_parse_iso_date(due_to, "due_to"))
        except ValidationError as exc:
            return JsonResponse({"detail": exc.message}, status=400)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tasks"
    headers = [
        "ID",
        "Title",
        "Description",
        "Status",
        "Priority",
        "Due Date",
        "Assigned To",
        "Created By",
        "Created At",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for task in tasks:
        worksheet.append(
            [
                task.id,
                task.title,
                task.description,
                task.status,
                task.priority,
                task.due_date,
                task.assigned_to.username if task.assigned_to else "",
                task.created_by.username,
                task.created_at.replace(tzinfo=None),
            ]
        )

    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 50)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="tasks_{timestamp}.xlsx"'
    return response


@login_required
@require_http_methods(["POST"])
def task_import_excel(request):
    enforce_tenant_access(request)
    upload = request.FILES.get("file")
    if not upload:
        return JsonResponse({"detail": "Missing file upload"}, status=400)
    status_code, payload = _import_tasks_from_upload(request, upload, request.POST.get("mapping"))
    return JsonResponse(payload, status=status_code)
