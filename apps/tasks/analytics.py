from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill


EXPECTED_COLUMNS = {
    "title": {"title", "task", "task_name", "name"},
    "description": {"description", "details", "notes"},
    "status": {"status", "state"},
    "priority": {"priority", "importance"},
    "due_date": {"due_date", "deadline", "due", "target_date"},
    "assigned_to": {"assigned_to", "owner", "assignee", "assigned"},
}


def _normalize(value: str) -> str:
    text = str(value).strip().lower()
    return "".join(ch if ch.isalnum() else "_" for ch in text).strip("_")


def _find_column(columns: list[str], candidates: set[str]) -> str | None:
    normalized = {_normalize(col): col for col in columns}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
    return None


def _write_table(worksheet, headers, rows):
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBFF", end_color="DDEBFF", fill_type="solid")
    for row in rows:
        worksheet.append(row)
    for column_cells in worksheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 50)


def analyze_task_dataframe(uploaded_file) -> tuple[dict, bytes]:
    dataframe = pd.read_excel(uploaded_file)
    if dataframe.empty:
        raise ValueError("Uploaded Excel file has no rows.")

    columns = list(dataframe.columns)
    mapped = {}
    for target, candidates in EXPECTED_COLUMNS.items():
        col = _find_column(columns, candidates)
        mapped[target] = col

    if not mapped["title"]:
        raise ValueError("Could not find a title column. Expected one of: title, task, task_name, name.")

    normalized_df = pd.DataFrame()
    normalized_df["title"] = dataframe[mapped["title"]].astype(str).str.strip()
    normalized_df["description"] = (
        dataframe[mapped["description"]].astype(str).str.strip() if mapped["description"] else ""
    )
    normalized_df["status"] = dataframe[mapped["status"]].astype(str).str.strip().str.lower() if mapped["status"] else "todo"
    normalized_df["priority"] = (
        dataframe[mapped["priority"]].astype(str).str.strip().str.lower() if mapped["priority"] else "medium"
    )
    normalized_df["assigned_to"] = (
        dataframe[mapped["assigned_to"]].astype(str).str.strip() if mapped["assigned_to"] else "unassigned"
    )
    normalized_df["due_date"] = pd.to_datetime(dataframe[mapped["due_date"]], errors="coerce") if mapped["due_date"] else pd.NaT

    status_map = {
        "todo": "todo",
        "to_do": "todo",
        "in_progress": "in_progress",
        "in progress": "in_progress",
        "done": "done",
        "completed": "done",
    }
    priority_map = {"low": "low", "medium": "medium", "med": "medium", "high": "high", "urgent": "high"}
    normalized_df["status"] = normalized_df["status"].map(lambda x: status_map.get(str(x).lower(), "todo"))
    normalized_df["priority"] = normalized_df["priority"].map(lambda x: priority_map.get(str(x).lower(), "medium"))
    normalized_df["assigned_to"] = normalized_df["assigned_to"].replace({"": "unassigned", "nan": "unassigned"})

    normalized_df["missing_title"] = normalized_df["title"] == ""
    normalized_df["duplicate_title"] = normalized_df["title"].duplicated(keep=False) & ~normalized_df["missing_title"]
    now = pd.Timestamp.utcnow().tz_localize(None)
    normalized_df["anomaly_due_date"] = (normalized_df["due_date"] < pd.Timestamp(year=2000, month=1, day=1)) | (
        normalized_df["due_date"] > now + pd.Timedelta(days=3650)
    )
    normalized_df["anomaly_due_date"] = normalized_df["anomaly_due_date"].fillna(False)
    normalized_df["anomaly_score"] = (
        normalized_df["missing_title"].astype(int)
        + normalized_df["duplicate_title"].astype(int)
        + normalized_df["anomaly_due_date"].astype(int)
    )

    cleaned_df = normalized_df.loc[~normalized_df["missing_title"]].copy()
    cleaned_df["due_date"] = cleaned_df["due_date"].dt.date

    status_priority_pivot = pd.pivot_table(
        cleaned_df, index="status", columns="priority", values="title", aggfunc="count", fill_value=0
    )
    assignee_status_pivot = pd.pivot_table(
        cleaned_df, index="assigned_to", columns="status", values="title", aggfunc="count", fill_value=0
    )

    summary = {
        "rows_uploaded": int(len(normalized_df)),
        "rows_after_cleaning": int(len(cleaned_df)),
        "rows_removed": int(len(normalized_df) - len(cleaned_df)),
        "duplicate_titles": int(normalized_df["duplicate_title"].sum()),
        "anomalous_due_dates": int(normalized_df["anomaly_due_date"].sum()),
        "generated_at": datetime.utcnow().isoformat(),
        "workflow_steps": [
            "Load Excel",
            "Normalize schema",
            "Clean and standardize values",
            "Detect anomalies",
            "Generate pivots and charts",
            "Build dashboard workbook",
        ],
    }

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    ws_flow = workbook.create_sheet("Workflow")
    _write_table(
        ws_flow,
        ["Step", "Status", "Description"],
        [[step, "completed", step] for step in summary["workflow_steps"]],
    )

    ws_quality = workbook.create_sheet("Data_Quality")
    _write_table(
        ws_quality,
        ["Metric", "Value"],
        [
            ["Rows Uploaded", summary["rows_uploaded"]],
            ["Rows After Cleaning", summary["rows_after_cleaning"]],
            ["Rows Removed", summary["rows_removed"]],
            ["Duplicate Titles", summary["duplicate_titles"]],
            ["Anomalous Due Dates", summary["anomalous_due_dates"]],
        ],
    )

    ws_raw = workbook.create_sheet("Raw_Normalized")
    raw_rows = normalized_df.fillna("").astype(str).values.tolist()
    _write_table(ws_raw, list(normalized_df.columns), raw_rows)

    ws_clean = workbook.create_sheet("Cleaned_Data")
    clean_rows = cleaned_df.fillna("").astype(str).values.tolist()
    _write_table(ws_clean, list(cleaned_df.columns), clean_rows)

    ws_pivot_1 = workbook.create_sheet("Pivot_Status_Priority")
    sp = status_priority_pivot.reset_index()
    _write_table(ws_pivot_1, list(sp.columns), sp.values.tolist())

    ws_pivot_2 = workbook.create_sheet("Pivot_Assignee_Status")
    aps = assignee_status_pivot.reset_index()
    _write_table(ws_pivot_2, list(aps.columns), aps.values.tolist())

    ws_dash = workbook.create_sheet("Dashboard")
    _write_table(
        ws_dash,
        ["KPI", "Value"],
        [
            ["Rows Uploaded", summary["rows_uploaded"]],
            ["Rows After Cleaning", summary["rows_after_cleaning"]],
            ["Rows Removed", summary["rows_removed"]],
            ["Duplicate Titles", summary["duplicate_titles"]],
            ["Anomalous Due Dates", summary["anomalous_due_dates"]],
        ],
    )
    ws_dash["D1"] = "Analyst Note"
    ws_dash["D2"] = "Pivot slicers are not generated by openpyxl; open Pivot sheets in Excel and insert slicers."
    ws_dash["D1"].font = Font(bold=True)

    if ws_pivot_1.max_row > 1 and ws_pivot_1.max_column > 1:
        chart = BarChart()
        chart.title = "Status by Priority"
        chart.y_axis.title = "Task Count"
        chart.x_axis.title = "Status"
        data_ref = Reference(ws_pivot_1, min_col=2, min_row=1, max_col=ws_pivot_1.max_column, max_row=ws_pivot_1.max_row)
        cats_ref = Reference(ws_pivot_1, min_col=1, min_row=2, max_row=ws_pivot_1.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 7
        chart.width = 12
        ws_dash.add_chart(chart, "A8")

    if ws_pivot_2.max_row > 1 and ws_pivot_2.max_column > 1:
        chart2 = BarChart()
        chart2.title = "Assignee Workload by Status"
        chart2.type = "bar"
        chart2.y_axis.title = "Assignee"
        chart2.x_axis.title = "Task Count"
        data_ref2 = Reference(
            ws_pivot_2,
            min_col=2,
            min_row=1,
            max_col=ws_pivot_2.max_column,
            max_row=ws_pivot_2.max_row,
        )
        cats_ref2 = Reference(ws_pivot_2, min_col=1, min_row=2, max_row=ws_pivot_2.max_row)
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cats_ref2)
        chart2.height = 8
        chart2.width = 12
        ws_dash.add_chart(chart2, "N8")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return summary, output.getvalue()
